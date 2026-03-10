import io
import csv
from datetime import datetime
from decimal import Decimal

from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import SalaryConfig, PayrollRecord, PayrollDeduction
from .serializers import SalaryConfigSerializer, PayrollRecordSerializer, PayrollDeductionSerializer
from .services import calculate_payroll
from apps.accounts.models import User
from apps.core.utils import get_request_company


def _is_accountant(user):
    """Check if user has payroll management access."""
    if user.is_staff:
        return True
    role_name = getattr(getattr(user, 'role', None), 'name', '') or ''
    return role_name.upper() in ('OWNER', 'ADMIN', 'ACCOUNTANT', 'HR')


class SalaryConfigViewSet(viewsets.ModelViewSet):
    serializer_class = SalaryConfigSerializer

    def get_queryset(self):
        company = get_request_company(self.request)
        if not company:
            return SalaryConfig.objects.none()
        return SalaryConfig.objects.filter(company=company, is_deleted=False).select_related('user', 'user__department')

    def perform_create(self, serializer):
        company = get_request_company(self.request)
        if not company:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Kompaniyangiz yo'q.")
        serializer.save(company=company)


class PayrollViewSet(viewsets.ModelViewSet):
    serializer_class = PayrollRecordSerializer

    def get_queryset(self):
        user = self.request.user
        company = get_request_company(self.request)
        if not company:
            return PayrollRecord.objects.none()

        qs = PayrollRecord.objects.filter(company=company, is_deleted=False).select_related(
            'user', 'user__department'
        ).prefetch_related('deduction_items')

        if not _is_accountant(user):
            qs = qs.filter(user=user)

        # Filters
        month = self.request.query_params.get('month')
        year = self.request.query_params.get('year')
        emp_id = self.request.query_params.get('user')
        dept_id = self.request.query_params.get('department')
        record_status = self.request.query_params.get('status')

        if month:
            try:
                dt = datetime.strptime(month, '%Y-%m')
                qs = qs.filter(month__year=dt.year, month__month=dt.month)
            except ValueError:
                pass
        if year:
            qs = qs.filter(month__year=year)
        if emp_id:
            qs = qs.filter(user_id=emp_id)
        if dept_id:
            qs = qs.filter(user__department_id=dept_id)
        if record_status:
            qs = qs.filter(status=record_status)

        return qs

    @action(detail=False, methods=['post'])
    def calculate(self, request):
        if not _is_accountant(request.user):
            return Response({'error': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)

        month_str = request.data.get('month')
        if not month_str:
            return Response({'error': 'month is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            month_date = datetime.strptime(month_str, '%Y-%m').date()
        except ValueError:
            return Response({'error': 'Invalid month format, use YYYY-MM'}, status=status.HTTP_400_BAD_REQUEST)

        users = User.objects.filter(company=request.user.company, is_active=True)
        count = 0
        skipped = 0
        for u in users:
            record = calculate_payroll(u, month_date)
            if record:
                count += 1
            else:
                skipped += 1

        return Response({
            'message': f'{count} ta xodim uchun maosh hisoblandi. {skipped} ta xodimda maosh sozlamasi yo\'q.',
            'calculated': count,
            'skipped': skipped,
        })

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        if not _is_accountant(request.user):
            return Response({'error': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        instance = self.get_object()
        instance.status = 'approved'
        instance.save()
        return Response({'status': 'approved', 'id': instance.id})

    @action(detail=True, methods=['post'], url_path='mark-paid')
    def mark_paid(self, request, pk=None):
        if not _is_accountant(request.user):
            return Response({'error': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        instance = self.get_object()
        instance.status = 'paid'
        instance.save()
        return Response({'status': 'paid', 'id': instance.id})

    @action(detail=False, methods=['post'], url_path='approve-bulk')
    def approve_bulk(self, request):
        if not _is_accountant(request.user):
            return Response({'error': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        ids = request.data.get('ids', [])
        month = request.data.get('month')
        company = get_request_company(request)
        qs = PayrollRecord.objects.filter(company=company, is_deleted=False)
        if ids:
            qs = qs.filter(id__in=ids)
        elif month:
            try:
                dt = datetime.strptime(month, '%Y-%m')
                qs = qs.filter(month__year=dt.year, month__month=dt.month, status='draft')
            except ValueError:
                return Response({'error': 'Invalid month'}, status=400)
        count = qs.update(status='approved')
        return Response({'approved': count})

    @action(detail=False, methods=['get'])
    def summary(self, request):
        if not _is_accountant(request.user):
            return Response({'error': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        company = get_request_company(request)
        qs = PayrollRecord.objects.filter(company=company, is_deleted=False)

        month = request.query_params.get('month')
        if month:
            try:
                dt = datetime.strptime(month, '%Y-%m')
                qs = qs.filter(month__year=dt.year, month__month=dt.month)
            except ValueError:
                pass

        agg = qs.aggregate(
            total_gross=Sum('base_salary') + Sum('overtime_pay') + Sum('night_pay'),
            total_net=Sum('net_salary'),
            total_tax=Sum('tax_amount'),
            total_inps=Sum('inps_amount'),
            total_deductions=Sum('deductions'),
            total_employees=Count('id'),
            paid_count=Count('id', filter=Q(status='paid')),
            approved_count=Count('id', filter=Q(status='approved')),
            draft_count=Count('id', filter=Q(status='draft')),
        )

        # Department breakdown
        dept_data = []
        dept_qs = qs.values('user__department__name').annotate(
            total_net=Sum('net_salary'),
            count=Count('id')
        ).order_by('-total_net')
        for d in dept_qs:
            dept_data.append({
                'department': d['user__department__name'] or 'Bo\'limsiz',
                'total_net': float(d['total_net'] or 0),
                'count': d['count'],
            })

        return Response({
            'total_gross': float(agg['total_gross'] or 0),
            'total_net': float(agg['total_net'] or 0),
            'total_tax': float(agg['total_tax'] or 0),
            'total_inps': float(agg['total_inps'] or 0),
            'total_deductions': float(agg['total_deductions'] or 0),
            'total_employees': agg['total_employees'],
            'paid_count': agg['paid_count'],
            'approved_count': agg['approved_count'],
            'draft_count': agg['draft_count'],
            'department_breakdown': dept_data,
        })

    @action(detail=False, methods=['get'])
    def my(self, request):
        qs = PayrollRecord.objects.filter(
            company=get_request_company(request),
            user=request.user,
            is_deleted=False
        ).prefetch_related('deduction_items')
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        if not _is_accountant(request.user):
            return Response({'error': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)

        qs = self.get_queryset()
        month = request.query_params.get('month', 'all')

        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="payroll_{month}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Xodim', 'Telefon', 'Bo\'lim', 'Oy',
            'Ish kunlari', 'Ish soatlari', 'Qo\'shimcha soatlar',
            'Asosiy maosh', 'Q.ish to\'lovi', 'Tunda ishlash',
            'Brutto', 'Daromad solig\'i', 'INPS', 'Boshqa chegirmalar',
            'Netto maosh', 'Holat', 'Eslatma',
        ])

        for rec in qs:
            gross = float(rec.base_salary) + float(rec.overtime_pay) + float(rec.night_pay)
            writer.writerow([
                rec.user.get_full_name() or rec.user.phone,
                rec.user.phone,
                getattr(getattr(rec.user, 'department', None), 'name', '') or '',
                rec.month.strftime('%Y-%m'),
                rec.work_days,
                float(rec.work_hours),
                float(rec.overtime_hours),
                float(rec.base_salary),
                float(rec.overtime_pay),
                float(rec.night_pay),
                gross,
                float(rec.tax_amount),
                float(rec.inps_amount),
                float(rec.deductions) - float(rec.tax_amount) - float(rec.inps_amount),
                float(rec.net_salary),
                rec.get_status_display(),
                rec.notes,
            ])

        return response

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        if not _is_accountant(request.user):
            return Response({'error': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)

        qs = self.get_queryset()
        month = request.query_params.get('month', 'all')

        wb = Workbook()
        ws = wb.active
        ws.title = f"Maosh {month}"

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0'),
        )
        money_fmt = '#,##0.00'
        alt_fill = PatternFill(start_color='F8F7FF', end_color='F8F7FF', fill_type='solid')

        headers = [
            '#', 'Xodim', 'Telefon', "Bo'lim", 'Oy',
            'Ish kunlari', 'Ish soatlari', "Q.soatlar",
            'Asosiy maosh', "Q.ish to'lovi", 'Tunda ishlash', 'Brutto',
            "Daromad solig'i", 'INPS', "Boshqa chegirmalar", 'Netto maosh',
            'Holat', 'Eslatma',
        ]
        col_widths = [5, 24, 16, 18, 10, 12, 12, 10, 16, 16, 14, 16, 16, 12, 16, 16, 12, 24]

        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = w

        ws.row_dimensions[1].height = 40

        status_labels = {'draft': 'Qoralama', 'approved': 'Tasdiqlangan', 'paid': 'To\'langan'}
        money_cols = [9, 10, 11, 12, 13, 14, 15, 16]

        for row_idx, rec in enumerate(qs, 2):
            gross = float(rec.base_salary) + float(rec.overtime_pay) + float(rec.night_pay)
            other_ded = float(rec.deductions) - float(rec.tax_amount) - float(rec.inps_amount)
            values = [
                row_idx - 1,
                rec.user.get_full_name() or rec.user.phone,
                rec.user.phone,
                getattr(getattr(rec.user, 'department', None), 'name', '') or '',
                rec.month.strftime('%Y-%m'),
                rec.work_days,
                float(rec.work_hours),
                float(rec.overtime_hours),
                float(rec.base_salary),
                float(rec.overtime_pay),
                float(rec.night_pay),
                gross,
                float(rec.tax_amount),
                float(rec.inps_amount),
                max(other_ded, 0),
                float(rec.net_salary),
                status_labels.get(rec.status, rec.status),
                rec.notes,
            ]
            fill = alt_fill if row_idx % 2 == 0 else None
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                if col in money_cols:
                    cell.number_format = money_fmt
                if fill:
                    cell.fill = fill

        # Totals row
        if qs.exists():
            total_row = qs.count() + 2
            ws.cell(row=total_row, column=1, value='JAMI').font = Font(bold=True)
            total_fill = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=total_row, column=col)
                cell.fill = total_fill
                cell.border = thin_border
                cell.font = Font(bold=True)

            agg = qs.aggregate(
                base=Sum('base_salary'), ot=Sum('overtime_pay'), night=Sum('night_pay'),
                tax=Sum('tax_amount'), inps=Sum('inps_amount'), ded=Sum('deductions'), net=Sum('net_salary')
            )
            totals = {
                9: float(agg['base'] or 0),
                10: float(agg['ot'] or 0),
                11: float(agg['night'] or 0),
                12: float((agg['base'] or 0) + (agg['ot'] or 0) + (agg['night'] or 0)),
                13: float(agg['tax'] or 0),
                14: float(agg['inps'] or 0),
                16: float(agg['net'] or 0),
            }
            for col, val in totals.items():
                cell = ws.cell(row=total_row, column=col, value=val)
                cell.number_format = money_fmt
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.border = thin_border

        ws.freeze_panes = 'B2'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="payroll_{month}.xlsx"'
        return response


class PayrollDeductionView(APIView):
    """Manage per-record deductions: list, create, delete."""

    def get(self, request, record_id):
        company = get_request_company(request)
        try:
            record = PayrollRecord.objects.get(id=record_id, company=company, is_deleted=False)
        except PayrollRecord.DoesNotExist:
            return Response({'error': 'Not found'}, status=404)
        items = PayrollDeduction.objects.filter(payroll_record=record, is_deleted=False)
        return Response(PayrollDeductionSerializer(items, many=True).data)

    def post(self, request, record_id):
        if not _is_accountant(request.user):
            return Response({'error': 'Not authorized'}, status=403)
        company = get_request_company(request)
        try:
            record = PayrollRecord.objects.get(id=record_id, company=company, is_deleted=False)
        except PayrollRecord.DoesNotExist:
            return Response({'error': 'Not found'}, status=404)

        ser = PayrollDeductionSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        item = ser.save(payroll_record=record)

        # Recalculate total deductions and net
        total_extra = PayrollDeduction.objects.filter(payroll_record=record, is_deleted=False).aggregate(s=Sum('amount'))['s'] or 0
        auto_ded = float(record.tax_amount) + float(record.inps_amount)
        record.deductions = round(auto_ded + float(total_extra), 2)
        gross = float(record.base_salary) + float(record.overtime_pay) + float(record.night_pay)
        record.net_salary = round(gross - float(record.deductions), 2)
        record.save(update_fields=['deductions', 'net_salary'])

        return Response(PayrollDeductionSerializer(item).data, status=201)

    def delete(self, request, record_id):
        if not _is_accountant(request.user):
            return Response({'error': 'Not authorized'}, status=403)
        company = get_request_company(request)
        item_id = request.query_params.get('item_id')
        if not item_id:
            return Response({'error': 'item_id required'}, status=400)
        try:
            record = PayrollRecord.objects.get(id=record_id, company=company, is_deleted=False)
            item = PayrollDeduction.objects.get(id=item_id, payroll_record=record, is_deleted=False)
        except (PayrollRecord.DoesNotExist, PayrollDeduction.DoesNotExist):
            return Response({'error': 'Not found'}, status=404)

        item.is_deleted = True
        item.save()

        # Recalculate
        total_extra = PayrollDeduction.objects.filter(payroll_record=record, is_deleted=False).aggregate(s=Sum('amount'))['s'] or 0
        auto_ded = float(record.tax_amount) + float(record.inps_amount)
        record.deductions = round(auto_ded + float(total_extra), 2)
        gross = float(record.base_salary) + float(record.overtime_pay) + float(record.night_pay)
        record.net_salary = round(gross - float(record.deductions), 2)
        record.save(update_fields=['deductions', 'net_salary'])

        return Response({'ok': True})
